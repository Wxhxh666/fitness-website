import csv
import io
import math
from datetime import datetime, date, timedelta
from urllib.parse import quote

from flask import jsonify, request, send_file, Response

from . import body_metrics_bp
from utils.auth import get_current_user, require_auth
from models.body_metric import BodyMetric
from models.body_record import BodyRecord, BodyGoal, BodyProfile
from models import db


# ---------- 常量与公式 ----------

ACTIVITY_MULTIPLIERS = {
    "sedentary": 1.2,
    "light": 1.375,
    "moderate": 1.55,
    "intense": 1.725,
    "athlete": 1.9,
}

ACTIVITY_LABELS = {
    "sedentary": "久坐少动",
    "light": "轻度运动",
    "moderate": "中度力量训练",
    "intense": "高强度训练",
    "athlete": "运动员级别",
}

STAGE_LABELS = {
    "fatloss": "减脂",
    "muscle": "增肌",
    "maintain": "保持",
}

GENDER_LABELS = {"male": "男", "female": "女"}


def _num(value):
    """安全转 float，None / 空 / 非法返回 None"""
    if value is None or value == "":
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    return v if math.isfinite(v) else None


def bmi_category(bmi):
    if bmi is None:
        return {"key": "unknown", "label": "未知"}
    if bmi < 18.5:
        return {"key": "underweight", "label": "偏瘦"}
    if bmi < 25:
        return {"key": "normal", "label": "正常范围"}
    if bmi < 30:
        return {"key": "overweight", "label": "超重"}
    return {"key": "obese", "label": "肥胖"}


def compute_body_fat(height_cm, weight_kg, gender, age, waist_cm=None, hip_cm=None, neck_cm=None):
    """体脂率：优先美国海军公式（需颈围），否则用 Deurenberg 简化公式"""
    gender = (gender or "male").lower()
    sex = 1 if gender == "male" else 0

    navy = None
    if sex == 1 and waist_cm and neck_cm and waist_cm > neck_cm > 0:
        try:
            navy = 495.0 / (1.0324 - 0.19077 * math.log10(waist_cm - neck_cm)
                            + 0.15456 * math.log10(height_cm)) - 450
        except (ValueError, ZeroDivisionError):
            navy = None
    elif sex == 0 and waist_cm and hip_cm and neck_cm and (waist_cm + hip_cm) > neck_cm > 0:
        try:
            navy = 495.0 / (1.29579 - 0.35004 * math.log10(waist_cm + hip_cm - neck_cm)
                            + 0.22100 * math.log10(height_cm)) - 450
        except (ValueError, ZeroDivisionError):
            navy = None

    if navy is not None and 3 < navy < 60:
        return round(navy, 1), "navy"

    if height_cm and weight_kg and age:
        bmi = weight_kg / ((height_cm / 100.0) ** 2)
        deurenberg = 1.2 * bmi + 0.23 * age - 10.8 * sex - 5.4
        return round(max(3.0, min(60.0, deurenberg)), 1), "deurenberg"
    return None, None


def compute_muscle_mass(weight_kg, height_cm, gender):
    if not weight_kg or not height_cm:
        return None
    if (gender or "male").lower() == "female":
        return round(0.184 * weight_kg + 0.277 * height_cm - 13.8, 1)
    return round(0.407 * weight_kg + 0.267 * height_cm - 19.2, 1)


def compute_metrics(data):
    """核心计算器：输入身体参数，输出全部指标 + 健康区间 + 健康提示"""
    height_cm = _num(data.get("height_cm"))
    weight_kg = _num(data.get("weight_kg"))
    waist_cm = _num(data.get("waist_cm"))
    hip_cm = _num(data.get("hip_cm"))
    neck_cm = _num(data.get("neck_cm"))
    age = _num(data.get("age"))
    gender = (data.get("gender") or "male").lower()
    activity = (data.get("activity_level") or "light").lower()

    result = {
        "height_cm": height_cm,
        "weight_kg": weight_kg,
        "waist_cm": waist_cm,
        "hip_cm": hip_cm,
        "gender": gender,
        "age": age,
        "activity_level": activity,
        "activity_label": ACTIVITY_LABELS.get(activity, activity),
        "bmi": None,
        "bmi_category": None,
        "bmi_category_label": None,
        "bmi_range": [18.5, 24.9],
        "body_fat": None,
        "body_fat_formula": None,
        "body_fat_range": [10, 20] if gender == "male" else [18, 28],
        "bmr": None,
        "bmr_range": [1000, 2500],
        "tdee": None,
        "tdee_multiplier": ACTIVITY_MULTIPLIERS.get(activity, 1.375),
        "tdee_range": [1200, 4500],
        "whr": None,
        "whr_range": [0.6, 1.1],
        "whr_healthy_max": 0.9 if gender == "male" else 0.85,
        "muscle_mass": None,
        "muscle_mass_range": None,
        "standard_weight": None,
        "healthy_weight_range": None,
        "health_tips": [],
    }

    if height_cm and weight_kg:
        height_m = height_cm / 100.0
        bmi = weight_kg / (height_m * height_m)
        result["bmi"] = round(bmi, 1)
        cat = bmi_category(result["bmi"])
        result["bmi_category"] = cat["key"]
        result["bmi_category_label"] = cat["label"]

        std_low = round(18.5 * height_m * height_m, 1)
        std_high = round(24.9 * height_m * height_m, 1)
        result["standard_weight"] = round(22 * height_m * height_m, 1)
        result["healthy_weight_range"] = [std_low, std_high]

        bf, formula = compute_body_fat(height_cm, weight_kg, gender, age,
                                       waist_cm, hip_cm, neck_cm)
        result["body_fat"] = bf
        result["body_fat_formula"] = formula

        if age:
            if gender == "male":
                bmr = 10 * weight_kg + 6.25 * height_cm - 5 * age + 5
            else:
                bmr = 10 * weight_kg + 6.25 * height_cm - 5 * age - 161
            result["bmr"] = round(bmr)
            result["tdee"] = round(bmr * ACTIVITY_MULTIPLIERS.get(activity, 1.375))

        result["muscle_mass"] = compute_muscle_mass(weight_kg, height_cm, gender)
        result["muscle_mass_range"] = [round(weight_kg * 0.25, 1), round(weight_kg * 0.60, 1)]

    if waist_cm and hip_cm:
        result["whr"] = round(waist_cm / hip_cm, 2)

    result["health_tips"] = build_health_tips(result)
    return result


def build_health_tips(m):
    """基于 BMI / 体脂 / 腰围 / 年龄生成简短健康提示（AI 饮食建议的前置参考）"""
    tips = []
    gender = (m.get("gender") or "male").lower()

    bmi = m.get("bmi")
    if bmi is not None:
        if bmi < 18.5:
            tips.append({"level": "warn", "text": "偏瘦：建议增加热量摄入并配合力量训练，避免单纯增重。"})
        elif bmi < 25:
            tips.append({"level": "info", "text": "BMI 处于正常范围，继续保持当前的饮食与训练节奏。"})
        elif bmi < 30:
            tips.append({"level": "warn", "text": "超重：建议控制每日热量摄入并加入规律有氧运动。"})
        else:
            tips.append({"level": "danger", "text": "肥胖：建议严格控卡 + 有氧训练，并咨询专业人士制定方案。"})

    bf = m.get("body_fat")
    if bf is not None:
        low = 6 if gender == "male" else 12
        high = 25 if gender == "male" else 35
        if bf < low:
            tips.append({"level": "danger", "text": "体脂过低：注意避免过度减脂，保护内分泌与激素水平。"})
        elif bf < m.get("body_fat_range", [10, 20])[1]:
            tips.append({"level": "info", "text": "体脂率处于健康区间，减脂 / 增肌节奏合理。"})
        elif bf > high:
            tips.append({"level": "warn", "text": "体脂偏高：建议热量缺口配合有氧与力量训练双管齐下。"})

    waist = m.get("waist_cm")
    if waist is not None:
        threshold = 90 if gender == "male" else 85
        if waist > threshold:
            tips.append({"level": "warn", "text": f"腰围 {waist:.0f}cm 超过健康参考（男 90cm / 女 85cm），存在中心性肥胖风险。"})

    whr = m.get("whr")
    if whr is not None:
        max_whr = 0.9 if gender == "male" else 0.85
        if whr > max_whr:
            tips.append({"level": "warn", "text": f"腰臀比 {whr:.2f} 偏高，脂肪更易堆积在腹部，注意加强核心与有氧。"})

    age = m.get("age")
    if age is not None and age >= 45:
        tips.append({"level": "info", "text": "45 岁以上建议定期体检，关注血压、血糖与骨密度指标。"})

    if not tips:
        tips.append({"level": "info", "text": "数据完整，暂无特别提醒；坚持记录以持续追踪身体变化。"})
    return tips


def current_user_id():
    user = get_current_user()
    return user.id if user else 0


def _upsert_metric(user_id, key, label, value, unit):
    """同步 legacy body_metrics 快照行，计算 change / trend"""
    if value is None:
        return
    row = BodyMetric.query.filter_by(user_id=user_id, key=key).first()
    new_value = round(float(value), 1)
    if row:
        previous = row.value
        row.value = new_value
        row.change = round(new_value - previous, 1)
        row.trend = "up" if row.change > 0 else "down" if row.change < 0 else "flat"
        row.recorded_at = datetime.utcnow()
    else:
        row = BodyMetric(user_id=user_id, key=key, label=label, value=new_value,
                         unit=unit, change=0.0, trend="flat", recorded_at=datetime.utcnow())
        db.session.add(row)


# ---------- 原有接口（兼容） ----------


@body_metrics_bp.route("", methods=["GET"])
def get_metrics():
    user = get_current_user()
    user_id = user.id if user else 0
    keys = ["weight", "body_fat", "muscle_mass", "bmr"]
    qs = BodyMetric.query.filter(
        BodyMetric.user_id == user_id,
        BodyMetric.key.in_(keys)
    ).order_by(BodyMetric.id).all()
    return jsonify(code=0, msg="success", data=[m.to_dict() for m in qs])


@body_metrics_bp.route("/bmi", methods=["POST"])
def calculate_bmi():
    data = request.get_json(silent=True) or {}
    height_cm = _num(data.get("height_cm"))
    weight_kg = _num(data.get("weight_kg"))
    if not height_cm or not weight_kg:
        return jsonify(code=400, msg="请提供 height_cm 和 weight_kg"), 400

    height_m = height_cm / 100.0
    bmi = round(weight_kg / (height_m * height_m), 1)
    cat = bmi_category(bmi)

    return jsonify(code=0, msg="success", data={
        "bmi": bmi,
        "category": cat["key"],
        "category_label": cat["label"],
        "healthy_range": "18.5 – 24.9"
    })


@body_metrics_bp.route("/measurements", methods=["GET"])
def get_measurements():
    user = get_current_user()
    user_id = user.id if user else 0
    keys = ["chest", "waist", "hips", "arm", "thigh", "calf"]
    qs = BodyMetric.query.filter(
        BodyMetric.user_id == user_id,
        BodyMetric.key.in_(keys)
    ).order_by(BodyMetric.id).all()
    return jsonify(code=0, msg="success", data=[m.to_dict() for m in qs])


@body_metrics_bp.route("/measurements/<int:measurement_id>", methods=["PUT"])
@require_auth
def update_measurement(measurement_id):
    m = BodyMetric.query.get_or_404(measurement_id)
    data = request.get_json(silent=True) or {}
    new_value = data.get("value")
    if new_value is None:
        return jsonify(code=400, msg="请提供 value"), 400

    previous = m.value
    m.value = round(float(new_value), 1)
    m.change = round(m.value - previous, 1)
    m.trend = "up" if m.change > 0 else "down" if m.change < 0 else "flat"
    m.recorded_at = datetime.utcnow()
    db.session.commit()

    result = m.to_dict()
    result["previous_value"] = previous
    return jsonify(code=0, msg="success", data=result)


@body_metrics_bp.route("/history", methods=["GET"])
def get_history():
    metric_key = request.args.get("metric_key")
    days = request.args.get("days", 30, type=int)
    since = datetime.utcnow() - timedelta(days=days)

    q = BodyMetric.query.filter(BodyMetric.recorded_at >= since)
    if metric_key:
        q = q.filter_by(key=metric_key)

    rows = q.order_by(BodyMetric.key, BodyMetric.recorded_at).all()

    grouped = {}
    for r in rows:
        grouped.setdefault(r.key, []).append({
            "date": r.recorded_at.strftime("%Y-%m-%d") if isinstance(r.recorded_at, datetime) else str(r.recorded_at)[:10],
            "value": r.value,
        })

    data = [{"metric_key": k, "records": v} for k, v in grouped.items()]
    return jsonify(code=0, msg="success", data=data)


# ---------- 新增：多维指标计算 ----------


@body_metrics_bp.route("/calculate", methods=["POST"])
def calculate():
    data = request.get_json(silent=True) or {}
    if not _num(data.get("height_cm")) or not _num(data.get("weight_kg")):
        return jsonify(code=400, msg="请至少提供身高与体重"), 400
    result = compute_metrics(data)
    return jsonify(code=0, msg="success", data=result)


# ---------- 新增：身材记录存档 / 列表 / 对比 / 导出 ----------


@body_metrics_bp.route("/records", methods=["GET"])
def list_records():
    user_id = current_user_id()
    days = request.args.get("days", type=int)
    limit = min(request.args.get("limit", 200, type=int), 1000)

    q = BodyRecord.query.filter_by(user_id=user_id)
    if days:
        since = date.today() - timedelta(days=days)
        q = q.filter(BodyRecord.record_date >= since)
    items = q.order_by(BodyRecord.record_date.desc(), BodyRecord.id.desc()).limit(limit).all()
    return jsonify(code=0, msg="success", data={
        "items": [r.to_dict() for r in items],
        "total": len(items),
    })


@body_metrics_bp.route("/records", methods=["POST"])
def save_record():
    data = request.get_json(silent=True) or {}
    user_id = current_user_id()

    # 缺少性别 / 年龄 / 强度时回退到个人档案
    profile = BodyProfile.query.filter_by(user_id=user_id).first()
    if not data.get("gender") and profile:
        data["gender"] = profile.gender
    if not data.get("age") and profile:
        data["age"] = profile.age
    if not data.get("activity_level") and profile:
        data["activity_level"] = profile.activity_level

    computed = compute_metrics(data)
    record = BodyRecord(
        user_id=user_id,
        record_date=_parse_date(data.get("record_date")),
        stage=(data.get("stage") or None),
        gender=computed["gender"],
        age=int(computed["age"]) if computed["age"] else None,
        activity_level=computed["activity_level"],
        height_cm=computed["height_cm"],
        weight_kg=computed["weight_kg"],
        waist_cm=computed["waist_cm"],
        hip_cm=computed["hip_cm"],
        neck_cm=_num(data.get("neck_cm")),
        chest_cm=_num(data.get("chest_cm")),
        shoulder_cm=_num(data.get("shoulder_cm")),
        thigh_cm=_num(data.get("thigh_cm")),
        arm_cm=_num(data.get("arm_cm")),
        calf_cm=_num(data.get("calf_cm")),
        bmi=computed["bmi"],
        body_fat=computed["body_fat"],
        bmr=computed["bmr"],
        tdee=computed["tdee"],
        whr=computed["whr"],
        muscle_mass=computed["muscle_mass"],
        standard_weight=computed["standard_weight"],
    )
    if not record.height_cm and not record.weight_kg and not any(
        [record.waist_cm, record.hip_cm, record.chest_cm, record.shoulder_cm,
         record.thigh_cm, record.arm_cm, record.calf_cm]
    ):
        return jsonify(code=400, msg="没有可保存的身体数据"), 400

    db.session.add(record)

    # 同步核心指标与围度快照行（保持旧卡片数据一致）
    if record.weight_kg:
        _upsert_metric(user_id, "weight", "体重", record.weight_kg, "kg")
    if record.body_fat:
        _upsert_metric(user_id, "body_fat", "体脂率", record.body_fat, "%")
    if record.muscle_mass:
        _upsert_metric(user_id, "muscle_mass", "肌肉量", record.muscle_mass, "kg")
    if record.bmr:
        _upsert_metric(user_id, "bmr", "基础代谢", record.bmr, "kcal")

    measure_map = [
        (record.chest_cm, "chest", "胸围"),
        (record.waist_cm, "waist", "腰围"),
        (record.hip_cm, "hips", "臀围"),
        (record.arm_cm, "arm", "上臂围"),
        (record.thigh_cm, "thigh", "大腿围"),
        (record.calf_cm, "calf", "小腿围"),
        (record.shoulder_cm, "shoulder", "肩宽"),
    ]
    for value, key, label in measure_map:
        if value:
            _upsert_metric(user_id, key, label, value, "cm")

    db.session.commit()
    return jsonify(code=0, msg="保存成功", data=record.to_dict())


@body_metrics_bp.route("/records/compare", methods=["GET"])
def compare_records():
    user_id = current_user_id()
    a_id = request.args.get("a", type=int)
    b_id = request.args.get("b", type=int)
    if not a_id or not b_id or a_id == b_id:
        return jsonify(code=400, msg="请选择两条不同的记录"), 400

    a = BodyRecord.query.filter_by(id=a_id, user_id=user_id).first()
    b = BodyRecord.query.filter_by(id=b_id, user_id=user_id).first()
    if not a or not b:
        return jsonify(code=404, msg="记录不存在"), 404

    fields = [
        ("weight_kg", "体重", "kg"),
        ("bmi", "BMI", ""),
        ("body_fat", "体脂率", "%"),
        ("waist_cm", "腰围", "cm"),
        ("whr", "腰臀比", ""),
        ("muscle_mass", "肌肉量", "kg"),
        ("bmr", "基础代谢", "kcal"),
        ("tdee", "每日消耗", "kcal"),
    ]
    diffs = []
    for key, label, unit in fields:
        av, bv = getattr(a, key), getattr(b, key)
        if av is None or bv is None:
            continue
        diff = round(bv - av, 2)
        diffs.append({
            "key": key,
            "label": label,
            "unit": unit,
            "from": av,
            "to": bv,
            "diff": diff,
        })

    summary_parts = []
    if a.weight_kg is not None and b.weight_kg is not None:
        dw = b.weight_kg - a.weight_kg
        summary_parts.append(f"体重{'下降' if dw < 0 else '增加'} {abs(dw):.1f}kg")
    if a.bmi is not None and b.bmi is not None:
        dbmi = b.bmi - a.bmi
        summary_parts.append(f"BMI {'下降' if dbmi < 0 else '上升'} {abs(dbmi):.1f}")
    if a.body_fat is not None and b.body_fat is not None:
        dbf = b.body_fat - a.body_fat
        summary_parts.append(f"体脂率{'下降' if dbf < 0 else '上升'} {abs(dbf):.1f}%")
    if a.waist_cm is not None and b.waist_cm is not None:
        dw2 = b.waist_cm - a.waist_cm
        summary_parts.append(f"腰围{'减少' if dw2 < 0 else '增加'} {abs(dw2):.1f}cm")
    if a.muscle_mass is not None and b.muscle_mass is not None:
        dm = b.muscle_mass - a.muscle_mass
        summary_parts.append(f"肌肉量{'增加' if dm > 0 else '减少'} {abs(dm):.1f}kg")

    if summary_parts:
        improving = (
            (b.weight_kg is not None and b.weight_kg <= a.weight_kg) or
            (b.body_fat is not None and b.body_fat <= a.body_fat) or
            (b.muscle_mass is not None and b.muscle_mass >= a.muscle_mass)
        )
        direction = "继续保持当前节奏，向目标稳步前进。" if improving else "建议复盘训练与饮食，及时调整计划。"
        summary = f"{a.record_date.strftime('%Y-%m-%d')} → {b.record_date.strftime('%Y-%m-%d')}：{'，'.join(summary_parts)}。{direction}"
    else:
        summary = "两条记录暂无可用对比数据。"

    return jsonify(code=0, msg="success", data={
        "from": a.to_dict(),
        "to": b.to_dict(),
        "diffs": diffs,
        "summary": summary,
    })


@body_metrics_bp.route("/records/<int:record_id>", methods=["DELETE"])
def delete_record(record_id):
    user_id = current_user_id()
    record = BodyRecord.query.get(record_id)
    if not record or record.user_id != user_id:
        return jsonify(code=404, msg="记录不存在"), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify(code=0, msg="已删除")


@body_metrics_bp.route("/records", methods=["DELETE"])
def delete_all_records():
    user_id = current_user_id()
    BodyRecord.query.filter_by(user_id=user_id).delete()
    db.session.commit()
    return jsonify(code=0, msg="全部记录已清空")


@body_metrics_bp.route("/export", methods=["GET"])
def export_records():
    """导出全部身材记录为 Excel（xlsx，缺依赖时回退 CSV）"""
    user_id = current_user_id()
    rows = BodyRecord.query.filter_by(user_id=user_id).order_by(
        BodyRecord.record_date.asc(), BodyRecord.id.asc()).all()

    headers = ["日期", "阶段", "性别", "年龄", "运动强度", "身高(cm)", "体重(kg)", "腰围(cm)", "臀围(cm)",
               "胸围(cm)", "肩宽(cm)", "大腿围(cm)", "手臂围(cm)", "小腿围(cm)",
               "BMI", "体脂率(%)", "BMR(kcal)", "TDEE(kcal)", "腰臀比", "肌肉量(kg)", "标准体重(kg)"]

    def row_values(r):
        return [
            r.record_date.strftime("%Y-%m-%d") if r.record_date else "",
            STAGE_LABELS.get(r.stage, r.stage or ""),
            GENDER_LABELS.get(r.gender, r.gender or ""),
            r.age, ACTIVITY_LABELS.get(r.activity_level, r.activity_level or ""),
            r.height_cm, r.weight_kg, r.waist_cm, r.hip_cm, r.chest_cm, r.shoulder_cm,
            r.thigh_cm, r.arm_cm, r.calf_cm, r.bmi, r.body_fat, r.bmr, r.tdee,
            r.whr, r.muscle_mass, r.standard_weight,
        ]

    filename = f"身材记录_{datetime.now().strftime('%Y%m%d')}"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "身材记录"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="111111")
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(row_values(r))
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value is not None else 8 for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 9), 22)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, as_attachment=True, download_name=f"{filename}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except ImportError:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(row_values(r))
        csv_bytes = "\ufeff" + output.getvalue()
        return Response(
            csv_bytes, mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=records.csv; filename*=UTF-8''{quote(filename + '.csv')}"}
        )


# ---------- 新增：个人基础档案 ----------


@body_metrics_bp.route("/profile", methods=["GET"])
def get_profile():
    user_id = current_user_id()
    profile = BodyProfile.query.filter_by(user_id=user_id).first()
    if not profile:
        return jsonify(code=0, msg="success", data={
            "gender": "male", "age": 25, "activity_level": "light",
            "unit_system": "metric",
        })
    return jsonify(code=0, msg="success", data=profile.to_dict(exclude=["id", "created_at", "updated_at"]))


@body_metrics_bp.route("/profile", methods=["POST"])
def save_profile():
    data = request.get_json(silent=True) or {}
    user_id = current_user_id()
    profile = BodyProfile.query.filter_by(user_id=user_id).first()
    if not profile:
        profile = BodyProfile(user_id=user_id)
        db.session.add(profile)

    if data.get("gender") in ("male", "female"):
        profile.gender = data["gender"]
    if isinstance(data.get("age"), (int, float)) and 10 <= data["age"] <= 100:
        profile.age = int(data["age"])
    if data.get("activity_level") in ACTIVITY_MULTIPLIERS:
        profile.activity_level = data["activity_level"]
    if data.get("unit_system") in ("metric", "imperial"):
        profile.unit_system = data["unit_system"]

    db.session.commit()
    return jsonify(code=0, msg="档案已保存", data=profile.to_dict(exclude=["id", "created_at", "updated_at"]))


# ---------- 新增：目标身材设定 ----------


@body_metrics_bp.route("/goals", methods=["GET"])
def list_goals():
    user_id = current_user_id()
    goals = BodyGoal.query.filter_by(user_id=user_id).order_by(BodyGoal.id.desc()).all()
    return jsonify(code=0, msg="success", data={"items": [g.to_dict() for g in goals]})


@body_metrics_bp.route("/goals", methods=["POST"])
def save_goal():
    data = request.get_json(silent=True) or {}
    user_id = current_user_id()
    target_weight = _num(data.get("target_weight"))
    target_bmi = _num(data.get("target_bmi"))
    target_body_fat = _num(data.get("target_body_fat"))
    daily_calorie = _num(data.get("daily_calorie"))
    if target_weight is None and target_bmi is None and target_body_fat is None:
        return jsonify(code=400, msg="请至少设定一个目标"), 400

    # 同一用户只保留一个当前目标
    BodyGoal.query.filter_by(user_id=user_id).delete()
    goal = BodyGoal(
        user_id=user_id,
        target_weight=target_weight,
        target_bmi=target_bmi,
        target_body_fat=target_body_fat,
        daily_calorie=daily_calorie,
    )
    db.session.add(goal)
    db.session.commit()
    return jsonify(code=0, msg="目标已保存", data=goal.to_dict())


@body_metrics_bp.route("/goals/<int:goal_id>", methods=["DELETE"])
def delete_goal(goal_id):
    user_id = current_user_id()
    goal = BodyGoal.query.get(goal_id)
    if not goal or goal.user_id != user_id:
        return jsonify(code=404, msg="目标不存在"), 404
    db.session.delete(goal)
    db.session.commit()
    return jsonify(code=0, msg="已删除")


def _parse_date(value):
    if not value:
        return date.today()
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return date.today()



